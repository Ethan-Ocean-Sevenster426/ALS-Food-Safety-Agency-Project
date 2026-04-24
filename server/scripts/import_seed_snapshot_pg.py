#!/usr/bin/env python3
"""
import_seed_snapshot_pg.py

Wipes selected tables on the prod Postgres DB (epvs) and re-imports rows
from server/seed-data/snapshot.xlsx, preserving IDs.

The 'users' table is NEVER wiped or re-imported by default. The intent
is to keep live prod accounts intact while replacing transactional data.

WARNING - because users are kept but every EPV/invoice/log row that
references a user comes from the dev snapshot, those references will
point at user IDs that may not exist in prod. The script disables FK
checks for the import session (SET session_replication_role = replica),
so the inserts succeed but you may end up with orphaned FKs. Pass
--include-users to wipe and re-import users from the snapshot too.

Usage (run from the repo root or server/):
    pip install psycopg2-binary openpyxl python-dotenv bcrypt
    python3 server/scripts/import_seed_snapshot_pg.py --dry-run
    python3 server/scripts/import_seed_snapshot_pg.py            # prompts
    python3 server/scripts/import_seed_snapshot_pg.py --yes      # no prompt
    python3 server/scripts/import_seed_snapshot_pg.py --include-users --yes
"""

from __future__ import annotations

import argparse
import datetime
import os
import re
import subprocess
import sys
from pathlib import Path

try:
    import openpyxl
    import psycopg2
    import psycopg2.extras
    from dotenv import load_dotenv
except ImportError as e:
    sys.stderr.write(
        f"missing dependency: {e}\n"
        "install with: pip install psycopg2-binary openpyxl python-dotenv bcrypt\n"
    )
    sys.exit(1)

# Optional - only needed when --include-users is set
try:
    import bcrypt
except ImportError:
    bcrypt = None


SCRIPT_DIR = Path(__file__).resolve().parent
SERVER_DIR = SCRIPT_DIR.parent
SNAPSHOT_PATH = SERVER_DIR / "seed-data" / "snapshot.xlsx"
ENV_PATH = SERVER_DIR / ".env"


# Insert order: parents before children (FK respect).
# Users sits where it would normally sit; we filter it out at runtime if
# --include-users is not passed.
INSERT_ORDER = [
    "KPITargets",
    "SupportTicketCategories",
    "Users",
    "ConsolidatedMasterAbattoirDatabase",
    "Invitations",
    "EggProductionVerifications",
    "EPVAuditLog",
    "EPVAttachments",
    "EPVInvoices",
    "ClientAuditLog",
    "LoginLog",
    "EmailSendLog",
    "SupportTickets",
    "SupportTicketComments",
]


# Snapshot sheet name -> Postgres table name.
TABLE_MAP = {
    "Users": "users",
    "Invitations": "invitations",
    "ConsolidatedMasterAbattoirDatabase": "client_records",
    "EggProductionVerifications": "egg_production_verifications",
    "EPVAuditLog": "epv_audit_log",
    "EPVAttachments": "epv_attachments",
    "EPVInvoices": "epv_invoices",
    "ClientAuditLog": "client_audit_log",
    "LoginLog": "login_log",
    "EmailSendLog": "email_send_log",
    "KPITargets": "kpi_targets",
    "SupportTicketCategories": "support_ticket_categories",
    "SupportTickets": "support_tickets",
    "SupportTicketComments": "support_ticket_comments",
}


# Columns to coerce to BOOLEAN when value is 0/1/'true'/'false'.
BOOLEAN_COLS = {
    "is_active",
    "is_reconciled",
    "is_verified",
    "is_paid",
    "is_manual",
    "manual_inspection",
    "completed",
}


DEFAULT_PASSWORD = "Password@123"
SUPER_ADMIN_EMAIL = "anthony@epvs.com"
SUPER_ADMIN_PASSWORD = "StrongPassword123!"


_camel_re_1 = re.compile(r"(.)([A-Z][a-z]+)")
_camel_re_2 = re.compile(r"([a-z0-9])([A-Z])")


def to_snake(name: str) -> str:
    """PascalCase / camelCase -> snake_case."""
    s = _camel_re_1.sub(r"\1_\2", name)
    s = _camel_re_2.sub(r"\1_\2", s)
    return s.lower()


def coerce_value(col: str, value):
    """Cast snapshot cell to a Postgres-friendly Python value."""
    if value is None or value == "":
        return None
    if col in BOOLEAN_COLS:
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        s = str(value).strip().lower()
        if s in ("1", "true", "t", "yes", "y"):
            return True
        if s in ("0", "false", "f", "no", "n"):
            return False
        return None
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return value
    return value


def load_snapshot(path: Path) -> dict[str, list[dict]]:
    if not path.exists():
        raise SystemExit(f"snapshot not found at {path}")
    wb = openpyxl.load_workbook(filename=str(path), data_only=True, read_only=True)
    out: dict[str, list[dict]] = {}
    for sheet_name in INSERT_ORDER:
        if sheet_name not in wb.sheetnames:
            out[sheet_name] = []
            continue
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            out[sheet_name] = []
            continue
        snake_header = [to_snake(str(h)) if h else None for h in header]
        rows: list[dict] = []
        for raw in rows_iter:
            if all(v is None or v == "" for v in raw):
                continue
            row: dict = {}
            for col, val in zip(snake_header, raw):
                if col is None:
                    continue
                row[col] = coerce_value(col, val)
            rows.append(row)
        out[sheet_name] = rows
    return out


def regenerate_user_hashes(users: list[dict]):
    if bcrypt is None:
        raise SystemExit(
            "bcrypt not installed - required when --include-users is used.\n"
            "install with: pip install bcrypt"
        )
    default_hash = bcrypt.hashpw(DEFAULT_PASSWORD.encode(), bcrypt.gensalt(10)).decode()
    super_hash = bcrypt.hashpw(SUPER_ADMIN_PASSWORD.encode(), bcrypt.gensalt(10)).decode()
    for u in users:
        email = str(u.get("email") or "").lower()
        u["password_hash"] = super_hash if email == SUPER_ADMIN_EMAIL else default_hash


def get_db_config() -> dict:
    if ENV_PATH.exists():
        load_dotenv(ENV_PATH)
    cfg = {
        "host": os.getenv("DB_HOST"),
        "port": os.getenv("DB_PORT", "5432"),
        "dbname": os.getenv("DB_NAME"),
        "user": os.getenv("DB_USER"),
        "password": os.getenv("DB_PASSWORD"),
    }
    missing = [k for k, v in cfg.items() if v in (None, "")]
    if missing:
        raise SystemExit(f"missing env vars: {missing} (looked in {ENV_PATH})")
    cfg["port"] = int(cfg["port"])
    return cfg


def take_backup(cfg: dict, label: str) -> Path:
    backups = SERVER_DIR / "backups"
    backups.mkdir(exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out = backups / f"epvs_backup_{label}_{stamp}.sql"
    print(f"[backup] running pg_dump -> {out}")
    env = os.environ.copy()
    env["PGPASSWORD"] = cfg["password"]
    cmd = [
        "pg_dump",
        "-h", cfg["host"],
        "-p", str(cfg["port"]),
        "-U", cfg["user"],
        "-d", cfg["dbname"],
        "-f", str(out),
    ]
    subprocess.run(cmd, env=env, check=True)
    size_mb = out.stat().st_size / 1024 / 1024
    print(f"[backup] wrote {out} ({size_mb:.1f} MB)")
    return out


def column_exists(cur, table: str, column: str) -> bool:
    cur.execute(
        """
        SELECT 1 FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = %s AND column_name = %s
        """,
        (table, column),
    )
    return cur.fetchone() is not None


def filter_row_to_existing_columns(cur, table: str, row: dict) -> dict:
    """Drop snapshot columns that don't exist on the live table.

    Schema drift between dev (where snapshot was exported) and prod is
    common - simply ignore unknown columns rather than crashing the
    whole import.
    """
    cur.execute(
        """
        SELECT column_name FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = %s
        """,
        (table,),
    )
    live_cols = {r[0] for r in cur.fetchall()}
    return {k: v for k, v in row.items() if k in live_cols}


def wipe(cur, table: str):
    cur.execute(
        "SELECT 1 FROM information_schema.tables "
        "WHERE table_schema = 'public' AND table_name = %s",
        (table,),
    )
    if not cur.fetchone():
        print(f"  (skip wipe) {table}: table does not exist")
        return
    cur.execute(f'DELETE FROM "{table}"')
    print(f"  wiped {table}")


def reset_sequence(cur, table: str):
    """If the table has an `id` sequence, jump it past max(id)."""
    cur.execute(
        """
        SELECT pg_get_serial_sequence(%s, 'id')
        """,
        (table,),
    )
    seq = cur.fetchone()[0]
    if not seq:
        return
    cur.execute(f'SELECT COALESCE(MAX(id), 0) FROM "{table}"')
    max_id = cur.fetchone()[0]
    next_val = max(max_id, 1)
    # If max_id == 0 we want next nextval() to return 1, so is_called=false.
    is_called = max_id > 0
    cur.execute("SELECT setval(%s, %s, %s)", (seq, next_val, is_called))


def bulk_insert(cur, table: str, rows: list[dict]) -> int:
    if not rows:
        print(f"  {table}: 0 rows")
        return 0
    inserted = 0
    for row in rows:
        clean = filter_row_to_existing_columns(cur, table, row)
        if not clean:
            continue
        cols = list(clean.keys())
        col_sql = ",".join(f'"{c}"' for c in cols)
        placeholders = ",".join(["%s"] * len(cols))
        values = [clean[c] for c in cols]
        try:
            cur.execute(
                f'INSERT INTO "{table}" ({col_sql}) VALUES ({placeholders})',
                values,
            )
            inserted += 1
        except psycopg2.Error as e:
            sys.stderr.write(
                f"  [warn] {table}: skipping row id={row.get('id')!r}: {e.pgerror or e}\n"
            )
            # Roll back this single statement and continue. We are inside
            # a savepoint per-row to make this work.
            cur.execute("ROLLBACK TO SAVEPOINT row_sp")
        finally:
            cur.execute("RELEASE SAVEPOINT row_sp")
            cur.execute("SAVEPOINT row_sp")
    print(f"  {table}: {inserted}/{len(rows)} rows inserted")
    return inserted


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true",
                        help="parse snapshot and print plan, do nothing to DB")
    parser.add_argument("--yes", action="store_true",
                        help="skip the interactive y/N confirmation")
    parser.add_argument("--include-users", action="store_true",
                        help="also wipe and re-import the users table from snapshot")
    parser.add_argument("--no-backup", action="store_true",
                        help="skip pg_dump backup (NOT recommended)")
    parser.add_argument("--snapshot", type=Path, default=SNAPSHOT_PATH,
                        help=f"path to snapshot xlsx (default: {SNAPSHOT_PATH})")
    args = parser.parse_args()

    print(f"[snapshot] loading {args.snapshot}")
    data = load_snapshot(args.snapshot)
    for sheet, rows in data.items():
        print(f"  {sheet}: {len(rows)} rows in snapshot")

    if args.include_users and data["Users"]:
        print("\n[users] regenerating bcrypt hashes for snapshot users")
        regenerate_user_hashes(data["Users"])

    skip_users = not args.include_users
    insert_order = [t for t in INSERT_ORDER if not (skip_users and t == "Users")]
    delete_order = list(reversed(insert_order))

    if args.dry_run:
        print("\n[dry-run] would wipe (in order):")
        for t in delete_order:
            print(f"  - {TABLE_MAP[t]}")
        print("[dry-run] would insert (in order):")
        for t in insert_order:
            print(f"  - {TABLE_MAP[t]}: {len(data[t])} rows")
        if skip_users:
            print("[dry-run] users table would be left untouched")
        else:
            print("[dry-run] users table WOULD be wiped + re-imported")
        return

    cfg = get_db_config()
    print(f"\n[db] {cfg['user']}@{cfg['host']}:{cfg['port']}/{cfg['dbname']}")

    if not args.yes:
        msg = (
            "\nThis will WIPE and RE-IMPORT the listed tables on the above DB."
            f"\nUsers table will be {'REPLACED' if not skip_users else 'KEPT untouched'}."
            "\nType 'yes' to proceed: "
        )
        if input(msg).strip().lower() != "yes":
            print("aborted.")
            return

    if not args.no_backup:
        try:
            take_backup(cfg, "preimport")
        except (FileNotFoundError, subprocess.CalledProcessError) as e:
            sys.stderr.write(
                f"\n[backup] FAILED: {e}\n"
                "Re-run with --no-backup to proceed without one (NOT recommended).\n"
            )
            sys.exit(1)

    conn = psycopg2.connect(**cfg)
    conn.autocommit = False
    try:
        with conn.cursor() as cur:
            # Disable FK checks + triggers for the session so dangling
            # snapshot references don't block inserts.
            cur.execute("SET session_replication_role = replica")

            print("\n[wipe] deleting existing rows (children first)")
            for sheet in delete_order:
                wipe(cur, TABLE_MAP[sheet])

            print("\n[insert] importing snapshot rows (parents first)")
            cur.execute("SAVEPOINT row_sp")
            total = 0
            for sheet in insert_order:
                total += bulk_insert(cur, TABLE_MAP[sheet], data[sheet])
            cur.execute("RELEASE SAVEPOINT row_sp")

            print("\n[seq] resetting id sequences past max(id)")
            for sheet in insert_order:
                reset_sequence(cur, TABLE_MAP[sheet])

            cur.execute("SET session_replication_role = origin")

        conn.commit()
        print(f"\n[done] committed. {total} rows inserted total.")
        if not args.include_users:
            print("[done] users table left untouched.")
    except Exception:
        conn.rollback()
        sys.stderr.write("\n[error] rolled back the entire transaction.\n")
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
